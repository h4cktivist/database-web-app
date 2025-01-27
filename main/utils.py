import os

import openpyxl
from openpyxl.styles import Font, Alignment
from docx import Document
from docx.enum.section import WD_ORIENT

from django.utils.timezone import now
from django.conf import settings

from celery import current_app
from celery.result import AsyncResult
from celery import states


def adjust_columns(sheet):
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width


def export_report(queryset, report_type):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    doc = Document()
    section = doc.sections[0]
    section.orientation = WD_ORIENT.LANDSCAPE
    new_width, new_height = section.page_height, section.page_width
    section.page_width = new_width
    section.page_height = new_height

    if report_type == 'sales':
        headers = [
            "ID",
            "Дата",
            "Сотрудник",
            "Покупатель",
            "Стоимость",
            "Дата и время сеанса",
        ]
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        for row_num, sale in enumerate(queryset, 2):
            staff_name = f"{sale.get('staff__first_name', '')} {sale.get('staff__last_name', '')}" if sale.get(
                'staff__first_name') and sale.get('staff__last_name') else ''
            customer_name = f"{sale.get('customer__first_name', '')} {sale.get('customer__last_name', '')}" if sale.get(
                'customer__first_name') and sale.get('customer__last_name') else ''
            session_datetime = f"{sale.get('ticket__session__session_date', '')} {sale.get('ticket__session__session_time', '')}" if sale.get(
                'ticket__session__session_date') and sale.get('ticket__session__session_time') else ''

            sheet.cell(row=row_num, column=1, value=sale.get('sale_id'))
            sheet.cell(row=row_num, column=2,
                       value=sale.get('date', '').strftime('%Y-%m-%d') if sale.get('date') else sale.get('date', ''))
            sheet.cell(row=row_num, column=3, value=staff_name)
            sheet.cell(row=row_num, column=4, value=customer_name)
            sheet.cell(row=row_num, column=5, value=sale.get('ticket__price'))
            sheet.cell(row=row_num, column=6, value=session_datetime)

        filename = f"sales_report_{now().strftime('%Y%m%d_%H%M%S')}"

    elif report_type == 'staff':
        columns = ['ID', 'Имя', 'Фамилия', 'Отчество', 'Должность', 'Кол-во продаж']
        for col_num, column in enumerate(columns, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = column
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        for row_num, staff in enumerate(queryset, 2):
            sheet.cell(row=row_num, column=1, value=staff.get('staff_id'))
            sheet.cell(row=row_num, column=2, value=staff.get('first_name'))
            sheet.cell(row=row_num, column=3, value=staff.get('last_name'))
            sheet.cell(row=row_num, column=4, value=staff.get('middle_name', ''))
            sheet.cell(row=row_num, column=5, value=staff.get('position__title', ''))
            sheet.cell(row=row_num, column=6, value=staff.get('total_sales', 0))

        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column].width = adjusted_width

        filename = f"staff_report_{now().strftime('%Y%m%d_%H%M%S')}"

    elif report_type == 'movies':
        columns = ['ID', 'Название', 'Жанр', 'Длительность', 'Рейтинг', 'Кол-во билетов']
        for col_num, column in enumerate(columns, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = column
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        for row_num, movie in enumerate(queryset, 2):
            sheet.cell(row=row_num, column=1, value=movie.get('movie_id'))
            sheet.cell(row=row_num, column=2, value=movie.get('title'))
            sheet.cell(row=row_num, column=3, value=movie.get('genre'))
            sheet.cell(row=row_num, column=4, value=movie.get('duration'))
            sheet.cell(row=row_num, column=5, value=movie.get('rating'))
            sheet.cell(row=row_num, column=6, value=movie.get('total_tickets_sold', 0))

        filename = f"movies_report_{now().strftime('%Y%m%d_%H%M%S')}"

    adjust_columns(sheet)

    filepath = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    workbook.save(f'{filepath}.xlsx')

    return f'{filepath}.xlsx'


def get_task_info(task_id):
    async_result = AsyncResult(task_id, app=current_app)
    if async_result.status == states.PENDING:
        return "Ожидает выполнения"
    elif async_result.status == states.STARTED:
        return "Выполняется"
    elif async_result.status == states.SUCCESS:
        return "Выполнен"
    elif async_result.status == states.FAILURE:
        return "Ошибка"
    else:
        return "Неизвестный статус"


def get_task_result(task_id):
    async_result = AsyncResult(task_id, app=current_app)
    if async_result.ready():
        try:
            result = async_result.get()
            return result
        except Exception as e:
            return {"error": f"Ошибка при выполнении задачи: {e}"}
    else:
        return {"status": async_result.status}
